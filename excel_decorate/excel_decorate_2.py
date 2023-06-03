import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('./test.xlsx', data_only=True)
ws = wb["Sheet1"]

good_color = PatternFill(start_color='99ff99', end_color='99ff99', fill_type='solid')
normal_color = PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')
bad_color = PatternFill(start_color='ff9999', end_color='ff9999', fill_type='solid')

cells = ws['C':'C']

for cell in cells:
    try:
        value = int(cell.value)
        
        if value >= 10000:
            cell.fill = good_color
        elif value >= 1000:
            cell.fill = normal_color
        else:
            cell.fill = bad_color
    except: 
        pass

wb.save('test_result.xlsx')