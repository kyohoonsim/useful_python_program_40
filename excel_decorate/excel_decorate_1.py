import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('./test.xlsx', data_only=True)
ws = wb["Sheet1"]
color = PatternFill(start_color='99ff99', end_color='99ff99', fill_type='solid')
ws.cell(3, 3).fill = color
wb.save('test_result.xlsx')